from __future__ import unicode_literals
import pandas as pd
import numpy as np
from string import ascii_lowercase
import itertools
from xlwt import Workbook
import io
import time
import calendar
import os
import re
import openpyxl
from openpyxl.styles import Alignment


# Function to convert SAP XLS to Dataframe
def xls_to_df(filename):

    file1 = io.open(filename, "r", encoding="utf-16")
    data = file1.readlines()
    xldoc = Workbook()
    sheet = xldoc.add_sheet("Sheet1", cell_overwrite_ok=True)
    for i, row in enumerate(data):
        for j, val in enumerate(row.replace('\n', '').split('\t')):
            sheet.write(i, j, val)
    xldoc.save('myexcel1.xlsx')
    df = pd.read_excel('myexcel1.xlsx')
    os.remove('myexcel1.xlsx')
    
    # Get a->aa, ab etc as columns 
    def iter_all_strings():
        mylist = []
        for size in itertools.count(1):
            for s in itertools.product(ascii_lowercase, repeat=size):
                yield "".join(s)                        
    alpha_list=[]
    for s in iter_all_strings():
        alpha_list.append(s)
        if s == 'bb':
            break 
    df.columns = alpha_list[0:len(df.columns)]
    return df


# Details
def get_details(df):
    start = list(df.b).index('DocumentNo')
    end = list(df.b).index('Tx')
    details = df[start:end]
    col_names = details.iloc[0]
    stripped_cols = [str(el).strip() for el in col_names]       
    details.columns = stripped_cols
    details =details.reset_index()
    details = details[['DocumentNo', 'Pstng Date', 'Reference', 'Doc. Date',
                        'Typ', 'Vendor', 'Business PartnerName', 'Tx', 'Rate', 'Base amount', 'Input tax']]
    details = details[(details.DocumentNo.notnull())]
    details = details[1:]
    details['Rate'] = details.Rate.apply(lambda x: x[:-4]+ '%')
    details=details.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    pivot = pd.pivot_table(details, index='Tx', values=['Base amount', 'Input tax'], aggfunc=np.sum)
    return details


# Summary
def get_summary(df):
    start = list(df.b).index('Tx')
    end = list(df.c).index('Balance of all company codes')
    summary = df[start:end]
    col_names = summary.iloc[0]
    stripped_cols = [str(el).strip() for el in col_names]       
    summary.columns = stripped_cols
    summary = summary.reset_index()
    summary = summary[['Tx', 'Rate', 'Description', 'Tax acct', 'Gross amount', 'Tax base amount',
                      'Non-deductible', 'Input tax']]
    summary = summary[2:]
    summary=summary.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    summary = summary[summary.Tx.notnull()]
    summary['Rate'] = summary.Rate.apply(lambda x: x[:-4]+ '%')
    summary=summary.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    
    return summary

# Balance of all company codes
def get_balance_co_codes(df):
    start = list(df.c).index('Balance of all company codes')
    balance = df[start:]
    balance = balance[balance.d.notnull()]
    col_names = balance.iloc[0]
    stripped_cols = [str(el).strip() for el in col_names]       
    balance.columns = stripped_cols
    balance = balance.reset_index()
    balance=balance[['Curr.', 'CoCd', 'Trs', 'Tx', 'Description', 'Rate', 'Tax base amount', 'Deductible', 'Balance']]
    balance = balance[1:]
    balance['Rate'] = balance.Rate.apply(lambda x: str(x)[:-4]+ '%') 
    balance=balance.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    
    return balance

# sort 1-12
def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def create_excel_wb(xls_files):
    tax_name = os.getcwd().split('/')[-1].split(' - ')[1:][0]    
    writer = pd.ExcelWriter(f"{tax_name}.xlsx", engine='xlsxwriter')
    
    if xls_files is not None:
        for file in xls_files:
            df = xls_to_df(file)
            summary = get_summary(df)
            details = get_details(df)
            balance = get_balance_co_codes(df)
            month = calendar.month_name[int(details.iloc[0]['Pstng Date'].split('.')[1])]
            details.to_excel(writer, sheet_name=f"{month} Details", index=False)
            get_summary(df).to_excel(writer, sheet_name=f"{month} Summary", index=False)
            get_balance_co_codes(df).to_excel(writer, sheet_name=f"{month} Bal", index=False)
        
    writer.save()
   
    
def format_excel_wb(excel_wb):
    tax_name = os.getcwd().split('/')[-1].split(' - ')[1:][0]
    wb = openpyxl.load_workbook(excel_wb)
    sheet = wb.sheetnames[1]
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    side = openpyxl.styles.Side(border_style=None)
    no_border = openpyxl.styles.borders.Border(
        left=side, 
        right=side, 
        top=side, 
        bottom=side,
    )
    
    
    numeric_cols = ['Base amount', 'Input tax', 'Tax base amount', 'Non-deductible', 'Input tax', 'Balance', 'Deductible']
    short_cols = ['Typ', 'Tx', 'Rate', 'CoCd', 'Curr.']
    for sheet in wb.worksheets:
        cols = ['A', 'B', 'C', 'D', 'E', 'F','G', 'H', "I", 'J', 'K']
    
        for col in cols:
            col_name = sheet[col+"1"].value
        
            # Numeric Format
            if col_name in numeric_cols:
                sheet.column_dimensions[col].number_format = '#,##0.00'
    
            if col_name in short_cols:
                sheet.column_dimensions[col].width = 10
            else:
                sheet.column_dimensions[col].width = 23
    
        for row in sheet:
            for cell in row:
                cell.border = no_border
                cell.alignment = Alignment(horizontal='left')
                
        # Add Header

    i=0
    for sheet in wb.worksheets:
        i+=1 # increment Page number
        sheet.oddHeader.left.text =f'\n\n\n\n\n{address1} \n{address2}\n{address3}\nPartita IVA: {vat_number}'
        sheet.oddHeader.centre.text = f"{centre_header}\n\n{tax}"
        sheet.oddHeader.right.text = f"Pagina &P/2020"

        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = False
        sheet.print_title_rows = '1:4'
        sheet.print_area = 'A:K'


    for sheet in wb.worksheets:
        sheet.insert_rows(1)
        sheet.insert_rows(2)
        sheet.insert_rows(4)
        sheet_type = sheet.title.split(" ")[1]
        month = sheet.title.split(" ")[0]

        if sheet_type == 'Details':
            sheet['A1'].value= f"{month} Detail"

        if sheet_type == 'Summary':
            sheet['A1'].value= f"{month} Tax Code Summary"

        if sheet_type == 'Bal':
            sheet['A1'].value= f"{month} Balance of Tax Codes"
                
    excel_wb = wb.save(f"{tax_name}.xlsx")
    return excel_wb


# Enter details here
divisions = {
    'div no.':{
        'co_cd': '',
        'address1' : "",
        'address2' : "",
        'address3' : "",
        'vat_number' : ',
        'centre_header' : f""
    }
}


# Setup
tax = os.getcwd().split('/')[-1].split(' - ')[1:][0]
division = str(input('What Divsion is this for?'))
co_cd =  divisions[division]['co_cd']
address1 = divisions[division]['address1']
address2 = divisions[division]['address2']
address3 = divisions[division]['address3']
vat_number = divisions[division]['vat_number']
centre_header = divisions[division]['centre_header']




dir_list = os.listdir()
print("Reading SAP Downloads...")
time.sleep(1.5)
xls_files = list(filter(lambda x : x[-3:]=='xls' , dir_list))
xls_files = sorted(xls_files, key=natural_keys)
print("Creating Excel Workbook")
time.sleep(1.5)
excel_wb = create_excel_wb(xls_files)
create_excel_wb(xls_files)
print("Formatting Excel Workbook")
time.sleep(1.5)
format_excel_wb(f"{tax}.xlsx")

print("All Done !!!")

